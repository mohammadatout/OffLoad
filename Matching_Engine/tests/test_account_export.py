"""Tests for the server-side AE Allocation Excel export."""

import os
import tempfile

import pytest
from openpyxl import load_workbook

from account_export import (
    DETAIL_SHEET,
    PRIMARY_SHEET,
    parse_columns_param,
    resolve_columns,
    stream_accounts_workbook,
    write_accounts_workbook,
)
from cisco_store import ACCOUNT_FIELDS, import_accounts_csv
from db import transaction
from settings_store import DEFAULT_ALLOCATION_COLUMN_KEYS

REAL_HEADER = (
    "SAVM_ID,UNIFIED_ACCOUNT_NAME,UNIFIED_STATE,SALES LEVEL 1,SALES LEVEL 2,"
    "SALES LEVEL 3,SALES LEVEL 4,SALES LEVEL 5,SALES LEVEL 6,SAV_VERTICAL_TOP,"
    "NODE_SEGMENT,NODE_SUBSEGMENT,NOMINATED_OWNER_CEC,NOMINATED_OWNER_EMAIL,"
    "NOMINATED_OWNER_NAME,NOMINATED_OWNER_JOB_TITLE,SOURCE,SAV_NAME,SFDC_ACC_NAME,"
    "SFDC_SAVM_ID,SAV_PEOPLE,SFDC_ACC_OWNER_EMAIL,EXISTS_IN_SAV,EXISTS_IN_SFDC,"
    "GS_ALL_EMAILS,NODE_ID,EDWSF_UPDATE_DTM,GS_MAX_END_DATE,NOMINATED_OWNER_IN_GS,"
    "NOMINATED_OWNER_IN_SFDC,NOMINATED_OWNER_IN_SAV,NOMINATION_PRIORITY,"
    "NOMINATION_REASON,CONFIDENCE_LEVEL,CANDIDATE_RANK"
)


def _row(savm_id="700000001", unified_name="ALPHA HOLDINGS", state="TX", sl2="US COMMERCIAL"):
    return ",".join(
        [
            savm_id, unified_name, state, "Americas", sl2,
            "COMMERCIAL EAST AREA", "TRI-STATE COMMERCIAL OPERATION",
            "PHILADELPHIA REGION", "CEA_PHILADELPHIA 1", "RETAIL",
            "COMMERCIAL", "COM-FOCUS", "aeone", "ae.one@example.com", "AE One",
            "Account Executive", "SAV+SFDC", "ALPHA GROUP",
            f"{unified_name} INC", savm_id, "AE One | aeone",
            "ae.one@example.com", "YES", "YES", "abarrien", "200283765",
            "00:10.6", "7/25/2026", "1", "1", "0", "2",
            "GS + SFDC agreement", "HIGH", "1",
        ]
    )


def _csv(*rows: str) -> bytes:
    return ("\n".join([REAL_HEADER, *rows]) + "\n").encode("utf-8")


@pytest.fixture
def seeded(db):
    import_accounts_csv(
        db,
        _csv(
            _row(),
            _row(savm_id="700000002", unified_name="BETA UTILITIES", state="OK"),
            _row(
                savm_id="700000003",
                unified_name="GAMMA HEALTH",
                state="AR",
                sl2="GLOBAL ENTERPRISE SEGMENT",
            ),
        ),
        "export.csv",
        actor="admin",
    )
    return db


def _load(path):
    return load_workbook(path, read_only=False)


def test_resolve_columns_falls_back_to_defaults():
    assert resolve_columns(None) == list(DEFAULT_ALLOCATION_COLUMN_KEYS)
    assert resolve_columns([]) == list(DEFAULT_ALLOCATION_COLUMN_KEYS)
    # Unknown keys are dropped, not interpolated into the sheet.
    assert resolve_columns(["nope", "bogus"]) == list(DEFAULT_ALLOCATION_COLUMN_KEYS)


def test_resolve_columns_star_means_every_field():
    assert resolve_columns(["*"]) == list(ACCOUNT_FIELDS)


def test_resolve_columns_keeps_order_and_deduplicates():
    assert resolve_columns(["state", "savm_group_id", "state"]) == [
        "state",
        "savm_group_id",
    ]


def test_parse_columns_param():
    assert parse_columns_param(None) is None
    assert parse_columns_param("") is None
    assert parse_columns_param("state, savm_group_id ,") == ["state", "savm_group_id"]


def test_workbook_has_both_sheets_with_business_labels(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    summary = write_accounts_workbook(
        seeded, path, selected=["savm_group_id", "unified_account_name", "state"]
    )
    assert summary["rows"] == 3
    assert summary["labels"] == ["SAV ID", "Unified Acc. Name", "State"]

    book = _load(path)
    assert book.sheetnames == [PRIMARY_SHEET, DETAIL_SHEET]
    header = [cell.value for cell in book[PRIMARY_SHEET][1]]
    assert header == ["SAV ID", "Unified Acc. Name", "State"]


def test_primary_sheet_holds_exactly_the_selected_columns(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    write_accounts_workbook(seeded, path, selected=["state", "savm_group_id"])

    book = _load(path)
    rows = list(book[PRIMARY_SHEET].values)
    assert rows[0] == ("State", "SAV ID")
    assert len(rows) == 4  # header plus three accounts
    assert {row[1] for row in rows[1:]} == {"700000001", "700000002", "700000003"}


def test_detail_sheet_keeps_every_field_regardless_of_selection(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    write_accounts_workbook(seeded, path, selected=["savm_group_id"])

    book = _load(path)
    header = [cell.value for cell in book[DETAIL_SHEET][1]]
    assert header == list(ACCOUNT_FIELDS)
    # Narrowing the front sheet must not remove data from the file.
    assert "sfdc_account_name" in header
    assert "am_reason" in header
    assert len(header) > len(["savm_group_id"])


def test_composite_hierarchy_column_is_rendered(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    write_accounts_workbook(seeded, path, selected=["sales_hierarchy"])

    book = _load(path)
    values = [row[0] for row in list(book[PRIMARY_SHEET].values)[1:]]
    assert any("Americas" in value and "US COMMERCIAL" in value for value in values)


def test_export_respects_filters(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    summary = write_accounts_workbook(
        seeded, path, selected=["state"], sl2="GLOBAL ENTERPRISE SEGMENT"
    )
    assert summary["rows"] == 1

    book = _load(path)
    assert [row[0] for row in list(book[PRIMARY_SHEET].values)[1:]] == ["AR"]


def test_export_respects_search(seeded, tmp_path):
    path = str(tmp_path / "out.xlsx")
    summary = write_accounts_workbook(seeded, path, selected=["state"], search="GAMMA")
    assert summary["rows"] == 1


def test_export_excludes_inactive_unless_asked(seeded, tmp_path):
    with transaction(seeded):
        seeded.execute(
            "UPDATE cisco_accounts SET is_active = 0 WHERE savm_group_id = ?",
            ("700000003",),
        )

    active_only = write_accounts_workbook(
        seeded, str(tmp_path / "a.xlsx"), selected=["state"]
    )
    assert active_only["rows"] == 2

    everything = write_accounts_workbook(
        seeded, str(tmp_path / "b.xlsx"), selected=["state"], include_inactive=True
    )
    assert everything["rows"] == 3


def test_export_handles_an_empty_result_set(db, tmp_path):
    path = str(tmp_path / "empty.xlsx")
    summary = write_accounts_workbook(db, path, selected=["state"])
    assert summary["rows"] == 0

    book = _load(path)
    assert book.sheetnames == [PRIMARY_SHEET, DETAIL_SHEET]
    assert [cell.value for cell in book[PRIMARY_SHEET][1]] == ["State"]


def test_full_backup_is_one_sheet_holding_every_field(seeded, tmp_path):
    # The pre-purge backup asks for '*'. A detail sheet would be a second copy
    # of the same data, so it is skipped and the export makes a single pass.
    path = str(tmp_path / "backup.xlsx")
    summary = write_accounts_workbook(seeded, path, selected=["*"])

    assert summary["has_detail_sheet"] is False
    book = _load(path)
    assert book.sheetnames == [PRIMARY_SHEET]
    header = [cell.value for cell in book[PRIMARY_SHEET][1]]
    assert len(header) == len(ACCOUNT_FIELDS)
    assert "SAV ID" in header and "SFDC Name" in header


def test_narrowed_selection_still_gets_a_detail_sheet(seeded, tmp_path):
    path = str(tmp_path / "narrow.xlsx")
    summary = write_accounts_workbook(seeded, path, selected=["savm_group_id"])
    assert summary["has_detail_sheet"] is True
    assert _load(path).sheetnames == [PRIMARY_SHEET, DETAIL_SHEET]


def test_stream_yields_a_valid_workbook_and_cleans_up(seeded, tmp_path):
    before = {
        name for name in os.listdir(tempfile.gettempdir())
        if name.startswith("offload_allocation_")
    }

    chunks = list(stream_accounts_workbook(seeded, selected=["savm_group_id"]))
    assert len(chunks) > 0

    payload = b"".join(chunks)
    # A real xlsx is a zip archive.
    assert payload[:2] == b"PK"

    path = tmp_path / "streamed.xlsx"
    path.write_bytes(payload)
    book = _load(str(path))
    assert book.sheetnames == [PRIMARY_SHEET, DETAIL_SHEET]
    assert len(list(book[PRIMARY_SHEET].values)) == 4

    after = {
        name for name in os.listdir(tempfile.gettempdir())
        if name.startswith("offload_allocation_")
    }
    assert after == before, "the temp workbook was left behind"
